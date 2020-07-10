# coding = tuf-8
'''
 读取文档的3中方式
 1.openpyxl
 2.
'''

import openpyxl
from openpyxl.utils import get_column_letter,column_index_from_string

path = r'C:\Users\fan\Desktop\用例1.xlsx'
path1 = r"C:\Users\fan\Desktop\个人资料表.xlsx"
wb = openpyxl.load_workbook(path1)           # 打开文件
# 获得sheet 名单
# print(wb.sheetnames)
# for sheet in wb:                            # 使用遍历获取所有表单
#     print(sheet.title)
# mysheet = wb.create_sheet('mySheet')        #创建工作簿
# print(wb.sheetnames)
# sheet2 = wb.get_sheet_by_name('mySheet')
# wb.save(path)
# wb.close()
ws = wb.active                              # 选择激活的工作表
# print(ws)
# print(ws['A1'].value)
#
# c = ws['B1']
# print('Row{},Column {} is {}'.format(c.row, c.column, c.value))         # 获得B1单元格的2个方法
# print('Cell {} is {}'.format(c.coordinate, c.value))
#
# print(ws.cell(row=1, column=2))                                         # 根据行、列获取单元格
# print(ws.cell(row=1, column=2).value)                                   # 根据行、列获取单元格的值
#
# for i in range(1,10):                                                   # 遍历行或列获得整行或整列的数据
#     #print(i, ws.cell(row=2, column=i).value)
#     print(i, ws.cell(row=i, column=1).value)

# colC =ws['C']               # 获得某一列
# # print(colC)
# # row6 = ws[6]
# # print(row6)
# col_range = ws['B:C']       # 切片行获得数据
# row_range = ws[1:4]         # 切片获得数据
# print("--------列----------")
# for col in col_range:               # 遍历列
#     for cell in col:                # 遍历列
#         print(cell.value)           # 取值
#
# print('----------行----------')
# for row in row_range:               # 遍历行
#     for cell in row:                # 遍历行
#         print(cell.value)           # 取值

# for row in ws.iter_rows(min_row =1, max_row=2, max_col=2):  # 遍历最大行数长度
#     for cell in row:                                        # 遍历行数
#         print(cell.value)                                         # 取值
#
# print(tuple(ws.rows))

# cell_range = ws['A1:C3']                    # 根据列切片获得数据
maxcolumn = ws.max_column
print(ws.max_column)
maxrow = ws.max_row
print(ws.max_row)
def get_excel_value():

    get_value = []                                                      # 读取excel文档的所有内容
    for row in range(2,maxrow+1):
        for cell in range(1,maxcolumn+1):
            get_value.append(ws.cell(row=row, column= cell).value)
            # print("第{}用例数据：{}".format(row-1,ws.cell(row=row, column= cell).value))
        print(get_value)

def get_excel_value2():
    for row in ws.iter_rows(min_row=2,min_col=1):
        print([cell.value for cell in row])



# for execute_col in ws.iter_cols(min_row =2, max_row=maxrow+1, min_col=9,max_col=9):
#     print(execute_col.)

    #if col.value == "yes":

def get_excutre():                              # 获取是否执行的用例，以列表形式返回执行用例的行号
    execute_col = ws['I']
    exe = []
    # print(execute_col)
    for i in execute_col[1:]:
        # print(i.coordinate,i.value)
        if i.value == 'yes':
            exe.append(i.row)
    return exe
        # value=i.coordinate
        # print(value)
def get_value(exe):
    need_value=[]
    for needrow in (exe):
        for row in ws.iter_rows(min_row=needrow,max_row=needrow):
            for cell in row:
                a=need_value.append(cell.value)
        need_value.append(a)
    return need_value
def red_case():
    list_case =[]
    for case in get_excutre():
        dictc_case=dict(
            id = ws.cell(row=case,column=1).value,
            mode = ws.cell(row=case,column=2).value,
            case_name = ws.cell(row=case,column=3).value,
            api_name = ws.cell(row=case,column=4).value,
            url = ws.cell(row=case,column=5).value,
            method = ws.cell(row=case,column=6).value,
            head = ws.cell(row=case,column=7).value,
            parmer = ws.cell(row=case,column=8).value,
            cause_or_not = ws.cell(row=case,column=9).value,
            expected_result = ws.cell(row=case,column=10).value)
        list_case.append(dictc_case)

    return list_case


cause_col = ws["L"]
# for row in ws.iter_rows(min_row=2,max_row=2):
#     for cell in row:
#         print(cell.value)

# if __name__ == '__main__':
#     exe=get_excutre()
#     print(get_value(exe))